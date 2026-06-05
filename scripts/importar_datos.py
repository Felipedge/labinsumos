"""
Script de importación de datos a Firebase Firestore
Ejecutar con: python3 scripts/importar_datos.py

Requiere:
  pip install openpyxl firebase-admin

Credenciales: colocar serviceAccountKey.json en la raíz del proyecto
o definir la variable de entorno GOOGLE_APPLICATION_CREDENTIALS
"""

import os
import sys
import datetime
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

# ──────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
KEY_PATH = os.path.join(BASE_DIR, "serviceAccountKey.json")

EXCEL_ESTANDARES = os.path.join(BASE_DIR, "scripts/data/Estandares_1.xlsx")
EXCEL_REACTIVOS  = os.path.join(BASE_DIR, "scripts/data/Reactivos_1.xlsx")
EXCEL_PLACEBOS   = os.path.join(BASE_DIR, "scripts/data/Placebos_y_Api_1.xlsx")

# Modo dry-run: True = solo imprime, no escribe en Firebase
DRY_RUN = "--dry-run" in sys.argv

# ──────────────────────────────────────────────
# INICIALIZAR FIREBASE
# ──────────────────────────────────────────────
if not DRY_RUN:
    cred = credentials.Certificate(KEY_PATH)
    firebase_admin.initialize_app(cred)
    db = firestore.client()
else:
    db = None
    print(">>> MODO DRY-RUN: no se escribirá nada en Firebase <<<\n")


# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def to_date(val):
    """Convierte datetime de openpyxl o None a datetime.datetime (sin hora)."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(val, datetime.date):
        return datetime.datetime(val.year, val.month, val.day)
    return None


def clean(val):
    """Normaliza strings vacíos y N/A."""
    if val is None:
        return None
    s = str(val).strip()
    if s.upper() in ("N/A", "NA", "", "NONE"):
        return None
    return s


def bool_field(val):
    if val is None:
        return False
    return str(val).strip().upper() in ("TRUE", "SI", "SÍ", "YES", "1")


def map_estado_estandar(raw):
    """Normaliza los estados del Excel de estándares al esquema de la app."""
    m = {
        "EN USO":    "En uso",
        "CERRADO":   "Cerrado",
        "SIN STOCK": "Sin stock",
        "VENCIDO":   "Sin stock",   # vencido se calcula dinámicamente; guardamos Sin stock
    }
    return m.get(str(raw).strip().upper(), "Cerrado")


def map_estado_reactivo(raw):
    """Normaliza los estados del Excel de reactivos."""
    m = {
        "EN USO":   "En uso",
        "SELLADO":  "Cerrado",
        "TERMINADO": "Sin stock",
    }
    return m.get(str(raw).strip().upper(), "Cerrado")


def map_estado_placebo(stock_val, vigencia_val, obs_col13, obs_col17):
    """
    Determina el estado de un placebo según las columnas del Excel.
    stock (col9): SI/NO  |  vigencia (col5): VENCIDO/+60DÍAS/etc
    obs_col13: texto libre con 'De baja'  |  obs_col17: 'De baja' / 'Vencido'
    """
    baja_text = " ".join(filter(None, [str(obs_col13 or ""), str(obs_col17 or "")])).lower()
    vigencia  = str(vigencia_val or "").strip().upper()
    stock     = str(stock_val or "").strip().upper()

    if "de baja" in baja_text:
        return "Dado de baja"
    if vigencia == "VENCIDO":
        return "Sin stock"
    if stock in ("SI", "SÍ"):
        return "En uso"
    return "Sin stock"


def write_doc(collection, doc_id, data):
    """Escribe un documento en Firestore (o imprime en dry-run)."""
    if DRY_RUN:
        print(f"  [{collection}/{doc_id}] {data}")
        return
    db.collection(collection).document(doc_id).set(data)


# ──────────────────────────────────────────────
# 1. ESTÁNDARES
# ──────────────────────────────────────────────
def importar_estandares():
    print("\n=== IMPORTANDO ESTÁNDARES ===")
    wb = openpyxl.load_workbook(EXCEL_ESTANDARES, data_only=True)
    ws = wb["Hoja1"]

    # Columnas (1-indexed):
    # 1=Código  2=ID(Std)  3=INGRESO  4=N°CAS  5=CLIENTE  6=NOMBRE  7=LOTE
    # 8=Producto  9=Potencia  10=Vencimiento  11=Vigencia  12=Vigencia primario
    # 13=Fecha revisión  14=Análisis posibles  15=Relevo Cant  16=Estado
    # 17=Observaciones  18=SECTOR  19=N° VIAL  20=Almacenamiento
    # 21=COA  22=Ficha seguridad  23=¿Tiene re-test?  24=Fabricante
    # 25=Cantidad recibida (mg)  26=Cant. por análisis (mg)

    ok = err = 0
    for r in range(2, ws.max_row + 1):
        codigo = clean(ws.cell(r, 1).value)
        if not codigo:
            continue

        fabricante = clean(ws.cell(r, 24).value) or ""
        es_usp = fabricante.upper() == "USP"

        doc = {
            "codigo":           codigo,
            "idStd":            clean(ws.cell(r, 2).value),
            "fechaIngreso":     to_date(ws.cell(r, 3).value),
            "numeroCAS":        clean(ws.cell(r, 4).value),
            "cliente":          clean(ws.cell(r, 5).value),
            "nombre":           clean(ws.cell(r, 6).value),
            "lote":             clean(ws.cell(r, 7).value),
            "producto":         clean(ws.cell(r, 8).value),
            "potencia":         ws.cell(r, 9).value,
            "fechaVencimiento": to_date(ws.cell(r, 10).value),
            "vigencia":         clean(ws.cell(r, 11).value),
            "vigenciaPrimario": clean(ws.cell(r, 12).value),
            "fechaRevision":    to_date(ws.cell(r, 13).value),
            "analisisPosibles": ws.cell(r, 14).value,
            "relieveCantMg":    ws.cell(r, 15).value,
            "estado":           map_estado_estandar(ws.cell(r, 16).value),
            "observaciones":    clean(ws.cell(r, 17).value),
            "sector":           clean(ws.cell(r, 18).value),
            "numeroVial":       clean(ws.cell(r, 19).value),
            "almacenamiento":   clean(ws.cell(r, 20).value),
            "coa":              bool_field(ws.cell(r, 21).value),
            "fichaSeguridad":   bool_field(ws.cell(r, 22).value),
            "tieneRetest":      bool_field(ws.cell(r, 23).value),
            "fabricante":       fabricante,
            "esUSP":            es_usp,
            "cantidadRecibidaMg": ws.cell(r, 25).value,
            "cantPorAnalisisMg":  ws.cell(r, 26).value,
            "creadoEn":         firestore.SERVER_TIMESTAMP if not DRY_RUN else "SERVER_TIMESTAMP",
            "importado":        True,
        }

        # Limpiar None en campos numéricos
        for k in ("analisisPosibles", "relieveCantMg", "cantidadRecibidaMg", "cantPorAnalisisMg", "potencia"):
            if doc[k] is None:
                doc[k] = 0

        try:
            write_doc("estandares", codigo, doc)
            ok += 1
            if ok % 100 == 0:
                print(f"  {ok} estándares procesados...")
        except Exception as e:
            print(f"  ERROR fila {r} ({codigo}): {e}")
            err += 1

    print(f"  ✓ {ok} estándares importados, {err} errores")


# ──────────────────────────────────────────────
# 2. REACTIVOS
# ──────────────────────────────────────────────
def importar_reactivos():
    print("\n=== IMPORTANDO REACTIVOS ===")
    wb = openpyxl.load_workbook(EXCEL_REACTIVOS, data_only=True)
    ws = wb["Hoja1"]

    # Encabezados en fila 5, datos desde fila 6
    # Col: 1=#  2=Fecha/Hora Recepción  3=Código General  4=Nombre  5=Lote
    # 6=Fecha vencimiento  7=Código Único  8=Lugar almacenamiento  9=Estado
    # 10=conteo  11=N°CAS  12=Mes/Año Recepción  13=Envase N°  14=Fabricante
    # 15=Proveedor  16=Envase indica fecha  17=Condición recepción
    # 18=Temperatura recepción  19=Envase indica condición  20=Tipo reactivo
    # 21=Tipo envase  22=Presentación  23=Integridad envase
    # 24=Documentación respaldo  25=Reactivo adquirido para  26=Producto recepcionado por

    ok = err = 0
    for r in range(6, ws.max_row + 1):
        codigo_unico = clean(ws.cell(r, 7).value)
        if not codigo_unico:
            continue

        # El N°CAS a veces viene como fecha de Excel (bug del Excel); limpiarlo
        ncas_raw = ws.cell(r, 11).value
        ncas = None
        if ncas_raw and not isinstance(ncas_raw, datetime.datetime):
            ncas = clean(str(ncas_raw))

        doc = {
            "codigo":           codigo_unico,
            "codigoGeneral":    clean(ws.cell(r, 3).value),
            "nombre":           clean(ws.cell(r, 4).value),
            "lote":             clean(str(ws.cell(r, 5).value)) if ws.cell(r, 5).value else None,
            "fechaVencimiento": to_date(ws.cell(r, 6).value),
            "estado":           map_estado_reactivo(ws.cell(r, 9).value),
            "lugarAlmacenamiento": clean(ws.cell(r, 8).value),
            "numeroCAS":        ncas,
            "mesAnioRecepcion": to_date(ws.cell(r, 12).value),
            "envaseNumero":     clean(ws.cell(r, 13).value),
            "fabricante":       clean(ws.cell(r, 14).value),
            "proveedor":        clean(ws.cell(r, 15).value),
            "envaseFecha":      clean(ws.cell(r, 16).value),
            "condicionRecepcion": clean(ws.cell(r, 17).value),
            "temperaturaRecepcion": ws.cell(r, 18).value,
            "tipoReactivo":     clean(ws.cell(r, 20).value),
            "tipoEnvase":       clean(ws.cell(r, 21).value),
            "presentacion":     clean(ws.cell(r, 22).value),
            "integridadEnvase": clean(ws.cell(r, 23).value),
            "documentacionRespaldo": clean(ws.cell(r, 24).value),
            "adquiridoPara":    clean(ws.cell(r, 25).value),
            "recepcionadoPor":  clean(ws.cell(r, 26).value),
            "fechaIngreso":     to_date(ws.cell(r, 2).value),
            "creadoEn":         firestore.SERVER_TIMESTAMP if not DRY_RUN else "SERVER_TIMESTAMP",
            "importado":        True,
        }

        try:
            write_doc("reactivos", codigo_unico, doc)
            ok += 1
            if ok % 200 == 0:
                print(f"  {ok} reactivos procesados...")
        except Exception as e:
            print(f"  ERROR fila {r} ({codigo_unico}): {e}")
            err += 1

    print(f"  ✓ {ok} reactivos importados, {err} errores")


# ──────────────────────────────────────────────
# 3. PLACEBOS
# ──────────────────────────────────────────────
def importar_placebos():
    print("\n=== IMPORTANDO PLACEBOS ===")
    wb = openpyxl.load_workbook(EXCEL_PLACEBOS, data_only=True)
    ws = wb["Placebo"]

    # Columnas (1-indexed):
    # 1=oricon(código)  2=Fecha ingreso  3=Lote  4=Vence  5=Vigencia
    # 6=Nombre  7=Laboratorio  8=Ubicación  9=Stock(SI/NO)
    # 10=Número de viales  11=Cantidad necesaria revalidación
    # 12=Uso  13=Observaciones  14=Código  17=De baja/Vencido (legend)

    ok = err = 0
    for r in range(2, ws.max_row + 1):
        codigo_base = clean(ws.cell(r, 1).value)
        if not codigo_base:
            continue

        codigo_completo = clean(ws.cell(r, 14).value) or codigo_base

        stock_val   = ws.cell(r, 9).value
        vigencia    = ws.cell(r, 5).value
        obs_col13   = ws.cell(r, 13).value
        obs_col17   = ws.cell(r, 17).value
        estado      = map_estado_placebo(stock_val, vigencia, obs_col13, obs_col17)

        # Fecha vencimiento: col 4 (puede ser datetime o 'N/A')
        fecha_venc_raw = ws.cell(r, 4).value
        fecha_venc = to_date(fecha_venc_raw) if isinstance(fecha_venc_raw, (datetime.datetime, datetime.date)) else None

        doc = {
            "codigo":              codigo_completo,
            "codigoBase":          codigo_base,
            "productoReferencia":  clean(ws.cell(r, 6).value),
            "laboratorio":         clean(ws.cell(r, 7).value),
            "lote":                clean(str(ws.cell(r, 3).value)) if ws.cell(r, 3).value else None,
            "fechaIngreso":        to_date(ws.cell(r, 2).value),
            "fechaVencimiento":    fecha_venc,
            "vigencia":            clean(str(vigencia)) if vigencia else None,
            "ubicacion":           clean(ws.cell(r, 8).value),
            "tieneStock":          str(stock_val or "").strip().upper() in ("SI", "SÍ"),
            "numeroViales":        ws.cell(r, 10).value,
            "cantidadRevalidacion": ws.cell(r, 11).value,
            "uso":                 clean(ws.cell(r, 12).value),
            "observaciones":       clean(ws.cell(r, 13).value),
            "estado":              estado,
            "creadoEn":            firestore.SERVER_TIMESTAMP if not DRY_RUN else "SERVER_TIMESTAMP",
            "importado":           True,
        }

        # Limpiar None en numéricos
        for k in ("numeroViales", "cantidadRevalidacion"):
            if doc[k] is None:
                doc[k] = 0

        try:
            write_doc("placebos", codigo_completo, doc)
            ok += 1
            if ok % 100 == 0:
                print(f"  {ok} placebos procesados...")
        except Exception as e:
            print(f"  ERROR fila {r} ({codigo_base}): {e}")
            err += 1

    print(f"  ✓ {ok} placebos importados, {err} errores")


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
if __name__ == "__main__":
    missing = [p for p in [EXCEL_ESTANDARES, EXCEL_REACTIVOS, EXCEL_PLACEBOS] if not os.path.exists(p)]
    if missing:
        print("ERROR: faltan archivos Excel:")
        for m in missing:
            print(f"  {m}")
        print("\nColoca los archivos en scripts/data/ con estos nombres:")
        print("  Estandares_1.xlsx")
        print("  Reactivos_1.xlsx")
        print("  Placebos_y_Api_1.xlsx")
        sys.exit(1)

    if not DRY_RUN and not os.path.exists(KEY_PATH):
        print(f"ERROR: no se encontró {KEY_PATH}")
        print("Descarga las credenciales de Firebase (serviceAccountKey.json) y colócalas en la raíz del proyecto.")
        sys.exit(1)

    importar_estandares()
    importar_reactivos()
    importar_placebos()

    print("\n✅ Importación completada.")
